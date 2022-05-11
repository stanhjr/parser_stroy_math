from openpyxl.reader import excel
from openpyxl.styles import PatternFill


def set_price_to_book(address, price, file_path):
    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active
    try:
        sheet[address] = price
        wb.save(file_path)
        return True
    finally:
        wb.save(file_path)


async def get_link_epicenter(last_row: int, file_path):

    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active
    counter = 2
    liter = 'K'
    list_all_market = []

    while counter != last_row:
        if sheet[liter + str(counter)].value:
            one_market = sheet[liter + str(counter)].value, sheet[liter + str(counter)].coordinate,
            list_all_market.append(one_market)
        counter += 1
    return list_all_market


def get_all_link_market(last_row: int, file_path):

    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active
    counter = 2
    row_excel_book = 'DEFGHIJKLMNOP'
    list_all_market = []

    while counter != last_row:
        for liter in row_excel_book:
            one_market = sheet[liter + str(counter)].value, sheet[liter + str(counter)].coordinate, sheet[liter + str(counter)].coordinate[0]
            list_all_market.append(one_market)
        counter += 1
    return list_all_market




async def find_last_row_excel(file_path):
    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active
    row_excel_book = 'DEFGHIJKLMNOP'
    row = 2
    while True:
        row += 1
        for cell_number in range(row, row + 1):
            counter = 0
            for liter in row_excel_book:
                if sheet[liter + str(cell_number)].value is None:
                    counter += 1
                if counter == len(row_excel_book):
                    return int(sheet[liter + str(cell_number)].coordinate[1:]) * len(row_excel_book), int(sheet[liter + str(cell_number)].coordinate[1:])

def is_float(string: str):
    try:
        string = string.replace(',', '.')
        result = float(string)
        return result
    except Exception as e:
        return False


def cut_paste_cell(file_path, last_row):
    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active
    list_cell_number = [x for x in range(1, last_row + 1)]
    list_cell_number.reverse()
    row_excel_book = 'RQPONMLKJIHG'
    for num in list_cell_number:
        for idx in range(len(row_excel_book)):
            if idx > 0:
                cell_old = row_excel_book[idx] + str(num)
                cell_new = row_excel_book[idx-1] + str(num)
                sheet[cell_new] = sheet[cell_old].value

        # g_cell = 'G' + str(num)
        # r_cell = 'R' + str(num)
        # sheet[g_cell] = sheet[r_cell].value
        # sheet[r_cell] = ''
    sheet["G1"] = "dnipro.atlant-shop.com.ua opt"
    wb.save(file_path)


def cut_paste_atlant_shop(file_path, last_row):
    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active
    list_cell_number = [x for x in range(2, last_row + 1)]
    for num in list_cell_number:
        g_cell = 'G' + str(num)
        r_cell = 'R' + str(num)
        sheet[g_cell] = sheet[r_cell].value
        sheet[r_cell] = ''
    wb.save(file_path)


def markup_excel_file(file_path, last_row):
    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active

    try:
        for i in range(2, last_row + 1):
            row_excel_book = 'EFGHIJKLMNOPRQ'
            d_cell = 'D' + str(i)
            d_value = is_float(sheet[d_cell].value)

            if not d_value:
                continue

            for letter in row_excel_book:
                cell = letter + str(i)
                cell_value = is_float(sheet[cell].value)
                if letter == 'E' and not cell_value:
                    sheet[cell] = 'Оптовой цены нет'
                if not cell_value:
                    continue
                elif cell_value > d_value:
                    sheet[cell].fill = PatternFill(start_color="BEE8C4", fill_type="solid")
                elif cell_value < d_value:
                    sheet[cell].fill = PatternFill(start_color="F7C5C5", fill_type="solid")

        wb.save(file_path)
        return file_path
    finally:
        wb.save(file_path)
